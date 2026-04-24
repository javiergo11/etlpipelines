"""
extractor_core.py
=================
Shared extraction and output logic for the EXR Owner Financials workflow.

This module contains ALL the extraction functions and output writers.
Both the command-line script and the Streamlit webapp import from here.

NOTHING in this file does I/O to the console (no print, no input).
It returns data and status messages — the caller decides how to display them.
"""

import os
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook          # openpyxl reads and writes Excel .xlsx files
from openpyxl.styles import Font, Alignment, PatternFill  # used to bold headers and color-code cells


# ---------------------------------------------------------------------------
# COA MAPPER (optional — extractor continues without it if module is missing)
# ---------------------------------------------------------------------------

# Try to import the COA mapper. If coa_mapper.py is missing the extractor still
# runs — it just skips the COA mapping tabs. This lets the file work in minimal
# environments without crashing.
try:
    from coa_mapper import COAMapper
    _COA_MAPPER_AVAILABLE = True
except ImportError:
    _COA_MAPPER_AVAILABLE = False

# Mappings at or above this confidence score are auto-accepted (no review needed).
# Must match CONFIDENCE_AUTO_ACCEPT in coa_mapper.py so color-coding stays in sync.
_COA_AUTO_ACCEPT = 0.85


# ---------------------------------------------------------------------------
# CONFIGURATION — EXTRA SPACE (EXR)
# (edit these if your EXR files use different label text)
# ---------------------------------------------------------------------------

# The EXR sheet names all start with a known prefix followed by a property number,
# e.g. "Rolling IS 7214". We search by prefix rather than exact name so the
# property number doesn't need to be hardcoded.
SHEET_PREFIXES = {
    "rolling_is": "Rolling IS",
    "unit_rate":  "Unit Rate",
    "ops_sum":    "Ops Sum",
    "rent_roll":  "Rent Roll",
}

# The rolling IS extraction scans from the first row containing the start label
# down to (and including) the row containing the stop label.
# EXR starts one row earlier at "Average Sq. Ft. Occupancy"; PS starts at "Rental Income".
EXR_ROLLING_IS_START_LABEL = "Average Sq. Ft. Occupancy"
ROLLING_IS_START_LABEL     = "Rental Income"
ROLLING_IS_STOP_LABEL      = "Net Operating Income"

# The Unit Rate sheet is scanned for any row whose label matches one of these.
# Values are looked up in the cells immediately to the right of the label.
UNIT_RATE_LABELS = [
    "Units Available",
    "Units Rented",
    "Sq Ft Available",
    "Sq Ft Rented",
]

# The Ops Sum sheet is scanned for rows matching these labels exactly.
# Any label not found is logged as a WARNING but does not stop extraction.
OPS_SUM_LABELS = [
    "Rentals During Month",
    "Walk In Rentals",
    "NSC rentals",
    "Web Rentals",
    "Vacates During Month",
    "Net Rentals",
]

# The expected column names in the EXR Rent Roll sheet header row.
# We find the header row by counting how many of these appear — if 3 or more
# match, we treat that row as the header. This makes detection resilient to
# minor formatting differences between file versions.
RENT_ROLL_HEADERS = [
    "Tenant Account",
    "Unit #",
    "Move-In Date",
    "Rent Rate",
    "Street Rate",
    "Paid-Thru Date",
    "Status",
    "Size",
    "Type",
]


# ---------------------------------------------------------------------------
# HELPER FUNCTIONS — SHARED (all formats)
# ---------------------------------------------------------------------------

def clean_label(text):
    """Normalize a cell label: strip whitespace, lowercase."""
    if text is None:
        return ""
    return str(text).strip().lower()


def label_matches(cell_text, target):
    """Check if cell text starts with target label (case-insensitive)."""
    return clean_label(cell_text).startswith(clean_label(target))


def is_date_value(value):
    """Check if a value is a date (datetime, 'Mon YYYY', 'Mon-YYYY', or 'Mon-YY' string)."""
    if isinstance(value, datetime):
        return True
    if isinstance(value, str):
        # EXR format: "Feb 2025"
        if re.match(r"[A-Z][a-z]{2}\s+\d{4}", value):
            return True
        # PS format: "Feb-2025"
        if re.match(r"[A-Z][a-z]{2}-\d{4}", value):
            return True
        # CS format: "Feb-26" (2-digit year)
        if re.match(r"[A-Z][a-z]{2}-\d{2}$", value):
            return True
    return False


def format_date(value):
    """Convert a date cell to 'Feb 2025' format. Normalizes PS and CS hyphen formats."""
    if isinstance(value, datetime):
        return value.strftime("%b %Y")
    if value is None:
        return ""
    s = str(value).strip()
    # Normalize CS 2-digit year "Feb-26" -> "Feb 2026"
    # (century pivot: 2-digit years are assumed to be 20xx — accurate through 2099)
    m = re.match(r"^([A-Z][a-z]{2})-(\d{2})$", s)
    if m:
        return f"{m.group(1)} 20{m.group(2)}"
    # Normalize PS hyphen format "Feb-2025" -> "Feb 2025"
    s = re.sub(r"^([A-Z][a-z]{2})-(\d{4})$", r"\1 \2", s)
    return s


def is_zero_row(values):
    """Return True if every numeric value is 0 or None."""
    for v in values:
        if v is None:
            continue
        try:
            if float(v) != 0:
                return False
        except (ValueError, TypeError):
            continue
    return True


def make_safe_filename(text):
    """Remove invalid Windows filename characters, replace spaces with underscores."""
    safe = re.sub(r'[\\/:*?"<>|]', "", text)
    safe = safe.replace(" ", "_")
    safe = re.sub(r"_+", "_", safe)
    safe = safe.strip("_")
    return safe


def guess_property_name(filename):
    """
    Guess a default property name from the input filename.
    'Feb__2026_Owner_Financials_-_EXR_Chattanooga.xlsx' -> 'EXR Chattanooga'
    """
    name = os.path.splitext(filename)[0]
    name = name.replace("_", " ")
    name = re.sub(r" +", " ", name)
    if " - " in name:
        name = name.rsplit(" - ", 1)[1]
    elif " – " in name:
        name = name.rsplit(" – ", 1)[1]
    return name.strip()


def find_sheet_by_prefix(workbook, prefix):
    """Find first sheet starting with prefix. Returns (name, ws) or (None, None)."""
    for name in workbook.sheetnames:
        if name.startswith(prefix):
            return name, workbook[name]
    return None, None


def extract_property_number(sheet_name, prefix):
    """Pull property number from sheet name like 'Rolling IS 7214'."""
    remainder = sheet_name.replace(prefix, "").strip()
    digits = re.sub(r"[^0-9]", "", remainder)
    return digits if digits else "UNKNOWN"


def calculate_sq_ft(size_str):
    """Convert '10X13' to 130 (width * depth)."""
    if size_str is None:
        return None
    match = re.match(r"(\d+)\s*[Xx]\s*(\d+)", str(size_str).strip())
    if match:
        return int(match.group(1)) * int(match.group(2))
    return None


def parse_date_string(date_str):
    """Convert 'Feb 2025' to (month, year, datetime). Returns (None,None,None) on failure."""
    try:
        dt = datetime.strptime(date_str, "%b %Y")
        return dt.month, dt.year, dt
    except ValueError:
        return None, None, None


# ---------------------------------------------------------------------------
# EXTRACTION FUNCTIONS — EXTRA SPACE (EXR)
# ---------------------------------------------------------------------------

def extract_rolling_is(ws):
    """
    Extract income statement from Rental Income through Net Operating Income.
    Returns (dates, rows) or (None, None).
    """
    # Load every row of the sheet into a plain Python list of lists.
    # We work with a local copy so we can scan back and forth freely.
    # values_only=False gives us Cell objects; we pull .value from each one.
    all_rows = []
    for row in ws.iter_rows(values_only=False):
        all_rows.append([c.value for c in row])

    if not all_rows:
        return None, None

    # -----------------------------------------------------------------------
    # Step 1: Find the date header row.
    # We scan every row looking for one that contains 5 or more date values.
    # That threshold avoids false positives from rows with only one or two dates.
    # Once found, we walk right from the first date cell, collecting each
    # formatted month string until we hit a non-date or empty cell.
    # -----------------------------------------------------------------------
    date_row_idx = None
    date_start_col = None
    date_end_col = None
    dates = []

    for idx, row in enumerate(all_rows):
        date_count = 0
        first_date_col = None
        for col_idx, val in enumerate(row):
            if is_date_value(val):
                date_count += 1
                if first_date_col is None:
                    first_date_col = col_idx  # remember where dates start
        if date_count >= 5:
            date_row_idx   = idx
            date_start_col = first_date_col
            # Walk right collecting dates; stop at the first gap or non-date
            for col_idx in range(first_date_col, len(row)):
                val = row[col_idx]
                if is_date_value(val):
                    dates.append(format_date(val))
                    date_end_col = col_idx
                elif val is None:
                    break
                else:
                    break  # non-date, non-None value (e.g. a YTD total column)
            break  # stop scanning rows — we found our date header

    if date_row_idx is None:
        return None, None

    # -----------------------------------------------------------------------
    # Step 2: Find the label column and the "Average Sq. Ft. Occupancy" start row.
    # We scan up to 30 rows below the date header looking for the first cell
    # in the leftmost 5 columns that starts with EXR_ROLLING_IS_START_LABEL.
    # This tells us both which column holds account labels and where data begins.
    # -----------------------------------------------------------------------
    label_col = None
    start_row_idx = None
    for idx in range(date_row_idx + 1, min(date_row_idx + 30, len(all_rows))):
        row = all_rows[idx]
        for col_idx in range(min(5, len(row))):
            if label_matches(row[col_idx], EXR_ROLLING_IS_START_LABEL):
                label_col     = col_idx
                start_row_idx = idx
                break
        if label_col is not None:
            break

    if label_col is None:
        return dates, None

    # -----------------------------------------------------------------------
    # Step 3: Collect data rows from "Rental Income" through "Net Operating Income".
    # For each row we build a dict: {"label": "...", "values": [jan, feb, ...]}.
    # Rows where every value is 0 or None are skipped (zero rows add no value).
    # We stop (inclusive) when we reach the stop label.
    # -----------------------------------------------------------------------
    extracted_rows = []
    num_date_cols = len(dates)

    for idx in range(start_row_idx, len(all_rows)):
        row = all_rows[idx]
        label_val = row[label_col] if label_col < len(row) else None

        # Skip blank rows (section gaps, empty lines between groups)
        if label_val is None or str(label_val).strip() == "":
            continue

        label_text = str(label_val).strip()

        # Collect the monthly values aligned to the dates list
        values = []
        for d in range(num_date_cols):
            col = date_start_col + d
            val = row[col] if col < len(row) else None
            values.append(val)

        # Skip rows where every month is zero — they add no information
        if not is_zero_row(values):
            extracted_rows.append({"label": label_text, "values": values})

        # Stop once we've captured the Net Operating Income row
        if label_matches(label_text, ROLLING_IS_STOP_LABEL):
            break

    return dates, extracted_rows


def extract_unit_rate(ws):
    """Extract summary metrics from Unit Rate sheet. Returns a dict."""
    results = {}

    # Build a lookup dict: {lowercased label: original label}.
    # This lets us match "units available" in the cell against "Units Available"
    # in our target list without caring about case.
    targets = {clean_label(t): t for t in UNIT_RATE_LABELS}

    for row in ws.iter_rows(values_only=False):
        cells = [c.value for c in row]
        for col_idx, val in enumerate(cells):
            if clean_label(val) in targets:
                original_label = targets[clean_label(val)]
                # The value sits somewhere in the next few columns to the right.
                # We scan up to 5 cells looking for the first numeric one.
                for search_col in range(col_idx + 1, min(col_idx + 5, len(cells))):
                    candidate = cells[search_col]
                    if candidate is not None:
                        try:
                            results[original_label] = float(candidate)
                            break  # found it — move on to the next label
                        except (ValueError, TypeError):
                            continue  # cell had text, keep looking right
    return results


def extract_ops_sum(ws):
    """Extract rental activity rows. Returns (dates, rows) or (None, None)."""
    all_rows = []
    for row in ws.iter_rows(values_only=False):
        all_rows.append([c.value for c in row])

    if not all_rows:
        return None, None

    date_row_idx = None
    date_start_col = None
    dates = []

    for idx, row in enumerate(all_rows):
        date_count = 0
        first_date_col = None
        for col_idx, val in enumerate(row):
            if is_date_value(val):
                date_count += 1
                if first_date_col is None:
                    first_date_col = col_idx
        if date_count >= 5:
            date_row_idx = idx
            date_start_col = first_date_col
            for col_idx in range(first_date_col, len(row)):
                val = row[col_idx]
                if is_date_value(val):
                    dates.append(format_date(val))
                elif val is None:
                    break
                else:
                    break
            break

    if date_row_idx is None:
        return None, None

    # Same lookup pattern as extract_unit_rate: lowercased label -> original label.
    targets = {clean_label(t): t for t in OPS_SUM_LABELS}
    num_date_cols = len(dates)

    # Find which column holds the row labels by scanning up to 40 rows below
    # the date header for the first cell that matches a known Ops Sum label.
    label_col = None
    for idx in range(date_row_idx + 1, min(date_row_idx + 40, len(all_rows))):
        row = all_rows[idx]
        for col_idx in range(min(4, len(row))):
            if clean_label(row[col_idx]) in targets:
                label_col = col_idx
                break
        if label_col is not None:
            break

    if label_col is None:
        return dates, None

    # Collect only the rows whose labels appear in OPS_SUM_LABELS.
    # Unlike Rolling IS we don't have a clean start/stop — we just filter
    # by label name and skip anything that doesn't match.
    extracted_rows = []
    for idx in range(date_row_idx + 1, len(all_rows)):
        row = all_rows[idx]
        raw_label = row[label_col] if label_col < len(row) else None
        if raw_label is None:
            continue
        cell_clean = clean_label(raw_label)
        if cell_clean in targets:
            values = []
            for d in range(num_date_cols):
                col = date_start_col + d
                val = row[col] if col < len(row) else None
                values.append(val)
            # Store with the canonical label name (not the raw cell text)
            extracted_rows.append({"label": targets[cell_clean], "values": values})

    return dates, extracted_rows


def extract_rent_roll(ws):
    """Extract rent roll data. Returns (headers, data_rows) or (None, None)."""
    all_rows = []
    for row in ws.iter_rows(values_only=False):
        all_rows.append([c.value for c in row])

    if not all_rows:
        return None, None

    header_row_idx = None
    # col_map will store {header_name: column_index} so we can pull each
    # field by name instead of relying on a fixed column order.
    col_map = {}

    # Scan every row building a temporary lookup of its cell values.
    # If 3 or more of our expected headers appear in one row, it's the header row.
    # Using 3 (not all 9) means detection still works if some columns are missing
    # in a particular file version.
    for idx, row in enumerate(all_rows):
        row_labels = {clean_label(cell): col_idx for col_idx, cell in enumerate(row)
                      if cell is not None}
        matches = sum(1 for h in RENT_ROLL_HEADERS if clean_label(h) in row_labels)
        if matches >= 3:
            header_row_idx = idx
            # Record the column index for every header we found
            for expected in RENT_ROLL_HEADERS:
                key = clean_label(expected)
                if key in row_labels:
                    col_map[expected] = row_labels[key]
            break

    if header_row_idx is None:
        return None, None

    data_rows = []
    for idx in range(header_row_idx + 1, len(all_rows)):
        row = all_rows[idx]

        # Stop at the first row that has no data in any known column.
        # This handles rent rolls that don't have a clean "end" marker.
        has_data = False
        for header in RENT_ROLL_HEADERS:
            if header in col_map:
                col_idx = col_map[header]
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None and str(val).strip() != "":
                    has_data = True
                    break
        if not has_data:
            break  # hit an empty row — we're past the data

        # Build an ordered list of values aligned to RENT_ROLL_HEADERS.
        # If a column was not found in this file, we insert None as a placeholder
        # so downstream code can always use the same positional index.
        row_values = []
        for header in RENT_ROLL_HEADERS:
            if header in col_map:
                col_idx = col_map[header]
                val = row[col_idx] if col_idx < len(row) else None
                row_values.append(val)
            else:
                row_values.append(None)  # column not present in this file

        # Skip vacant and non-tenant units
        status_idx = RENT_ROLL_HEADERS.index("Status")
        if status_idx < len(row_values):
            status_val = str(row_values[status_idx] or "").strip().lower()
            if status_val in ("available", "company use"):
                continue

        # Calculate Sq Ft
        size_idx = RENT_ROLL_HEADERS.index("Size")
        sq_ft = calculate_sq_ft(row_values[size_idx]) if size_idx < len(row_values) else None
        row_values.append(sq_ft)

        data_rows.append(row_values)

    output_headers = RENT_ROLL_HEADERS + ["Sq Ft"]
    return output_headers, data_rows


# ---------------------------------------------------------------------------
# EXTRACTION FUNCTIONS — PUBLIC STORAGE (PS)
# ---------------------------------------------------------------------------

# Section header rows in the PS IS sheet that are labels only — no data values.
# These appear inside the extraction range and must be skipped entirely.
PS_SECTION_HEADERS = {
    "revenue",
    "contractually set fees",
    "other expenses",
    "other items",
}

# Rows that are PS-calculated subtotals — extracted but tagged PS_Rollup in COA mapping.
# Listed here so extract_ps_rolling_is can include them (not skip them).
PS_ROLLUP_LABELS = {
    "total revenue",
    "total contractually set fees",
    "total other expenses",
    "total operating expenses",
    "net operating income",
}


def extract_ps_property_number(ws):
    """
    Parse the PS property number from cell B3.
    'B3 = "77712 - Wentworth (Vacaville, CA)"' -> "77712"
    Returns empty string if the pattern is not found.
    """
    try:
        val = ws["B3"].value
        if val:
            match = re.match(r"(\d+)", str(val).strip())
            if match:
                return match.group(1)
    except Exception:
        pass
    return ""


def extract_ps_rolling_is(ws):
    """
    Extract the PS income statement from the IS sheet.

    PS-specific behaviour:
      - Date row is row 7 (index 6); dates are in columns C:N (indices 2–13)
      - Dates are formatted 'Feb-2025' — normalised to 'Feb 2025' by format_date()
      - Column O (index 14) contains a YTD total — excluded automatically because
        it does not pass is_date_value() after normalisation check
      - Labels are in column B (index 1)
      - Section header rows (Revenue, Contractually set fees, etc.) are skipped
      - Zero rows are dropped
      - Extraction stops (inclusive) at 'Net Operating Income'

    Returns (dates, rows) or (None, None).
    """
    all_rows = []
    for row in ws.iter_rows(values_only=False):
        all_rows.append([c.value for c in row])

    if not all_rows:
        return None, None

    # Find the date header row — same algorithm as extract_rolling_is()
    date_row_idx  = None
    date_start_col = None
    date_end_col   = None
    dates = []

    for idx, row in enumerate(all_rows):
        date_count     = 0
        first_date_col = None
        for col_idx, val in enumerate(row):
            if is_date_value(val):
                date_count += 1
                if first_date_col is None:
                    first_date_col = col_idx
        if date_count >= 5:
            date_row_idx   = idx
            date_start_col = first_date_col
            for col_idx in range(first_date_col, len(row)):
                val = row[col_idx]
                if is_date_value(val):
                    dates.append(format_date(val))
                    date_end_col = col_idx
                elif val is None:
                    break
                else:
                    break
            break

    if date_row_idx is None:
        return None, None

    # Find the label column and the start row (first row with "Rental Income")
    label_col     = None
    start_row_idx = None
    for idx in range(date_row_idx + 1, min(date_row_idx + 30, len(all_rows))):
        row = all_rows[idx]
        for col_idx in range(min(5, len(row))):
            if label_matches(row[col_idx], ROLLING_IS_START_LABEL):
                label_col     = col_idx
                start_row_idx = idx
                break
        if label_col is not None:
            break

    if label_col is None:
        return dates, None

    # Collect rows, skipping PS section headers and zero rows
    extracted_rows = []
    num_date_cols  = len(dates)

    for idx in range(start_row_idx, len(all_rows)):
        row       = all_rows[idx]
        label_val = row[label_col] if label_col < len(row) else None

        if label_val is None or str(label_val).strip() == "":
            continue

        label_text  = str(label_val).strip()
        label_lower = label_text.lower()

        # Skip pure section header rows — they carry no numeric data
        if label_lower in PS_SECTION_HEADERS:
            continue

        values = []
        for d in range(num_date_cols):
            col = date_start_col + d
            val = row[col] if col < len(row) else None
            values.append(val)

        if not is_zero_row(values):
            extracted_rows.append({"label": label_text, "values": values})

        if label_matches(label_text, ROLLING_IS_STOP_LABEL):
            break

    return dates, extracted_rows


def extract_ps_rent_roll_occupancy(ws):
    """
    Count occupied units from the PS Rent Roll sheet.

    PS rent rolls do not carry the rich tenant data that EXR provides.
    The only reliable metric is the count of active rows: column C (Account #)
    starting at row 8 (index 7), stopping at the first empty cell.

    Returns an int (occupied unit count), or None if the sheet is unusable.
    """
    try:
        count = 0
        for row in ws.iter_rows(min_row=8, min_col=3, max_col=3, values_only=True):
            val = row[0]
            if val is None or str(val).strip() == "":
                break
            count += 1
        return count if count > 0 else None
    except Exception:
        return None


# ---------------------------------------------------------------------------
# EXTRACTION FUNCTIONS — CUBESMART (CS)
# ---------------------------------------------------------------------------

# Sheet name is exact ("Rolling Details"), unlike EXR which uses a prefix+number.
CS_ROLLING_IS_SHEET       = "Rolling Details"
CS_ROLLING_IS_START_LABEL = "Rental Income"
# Stop label: reuse ROLLING_IS_STOP_LABEL = "Net Operating Income".
# label_matches() uses startswith so it also matches "Net Operating Income (Loss)".


def extract_cs_property_number(ws):
    """
    Parse the CS property number from cell O1.
    'O1 = "3534 CUBESMART AR LITTLE ROCK PRATT RD"' -> "3534"
    Returns empty string if the pattern is not found.
    """
    try:
        val = ws["O1"].value
        if val:
            match = re.match(r"(\d+)", str(val).strip())
            if match:
                return match.group(1)
    except Exception:
        pass
    return ""


def extract_cs_rolling_is(ws):
    """
    Extract the CS income statement from the Rolling Details sheet.

    CS-specific behaviour:
      - Dates formatted 'Feb-26' (2-digit year) — normalised to 'Feb 2026' by format_date()
      - '12 Month Total' column sits immediately after the last month — excluded
        automatically because the header string doesn't pass is_date_value()
      - Labels in column B (same as PS)
      - Extraction stops (inclusive) at 'Net Operating Income (Loss)'
      - After extraction, any month column whose NOI value is 0 or None is dropped
        from both the dates list and every row's values list. This removes empty
        future months in a rolling 12-month view.

    Returns (dates, rows) or (None, None).
    """
    all_rows = []
    for row in ws.iter_rows(values_only=False):
        all_rows.append([c.value for c in row])

    if not all_rows:
        return None, None

    # Find the date header row — same algorithm as extract_rolling_is()
    date_row_idx   = None
    date_start_col = None
    dates          = []

    for idx, row in enumerate(all_rows):
        date_count     = 0
        first_date_col = None
        for col_idx, val in enumerate(row):
            if is_date_value(val):
                date_count += 1
                if first_date_col is None:
                    first_date_col = col_idx
        if date_count >= 5:
            date_row_idx   = idx
            date_start_col = first_date_col
            for col_idx in range(first_date_col, len(row)):
                val = row[col_idx]
                if is_date_value(val):
                    dates.append(format_date(val))
                elif val is None:
                    break
                else:
                    break  # non-date (e.g. "12 Month Total") ends the date walk
            break

    if date_row_idx is None:
        return None, None

    # Find the label column and the start row ("Rental Income")
    label_col     = None
    start_row_idx = None
    for idx in range(date_row_idx + 1, min(date_row_idx + 30, len(all_rows))):
        row = all_rows[idx]
        for col_idx in range(min(5, len(row))):
            if label_matches(row[col_idx], CS_ROLLING_IS_START_LABEL):
                label_col     = col_idx
                start_row_idx = idx
                break
        if label_col is not None:
            break

    if label_col is None:
        return dates, None

    # Collect rows through "Net Operating Income (Loss)" (inclusive)
    extracted_rows = []
    num_date_cols  = len(dates)

    for idx in range(start_row_idx, len(all_rows)):
        row       = all_rows[idx]
        label_val = row[label_col] if label_col < len(row) else None

        if label_val is None or str(label_val).strip() == "":
            continue

        label_text = str(label_val).strip()

        values = []
        for d in range(num_date_cols):
            col = date_start_col + d
            val = row[col] if col < len(row) else None
            values.append(val)

        # Existing per-value zero check: a row stays in if ANY value is non-zero.
        # This correctly keeps rows like "Property Taxes" where monthly values
        # offset to a zero sum but individual months carry data.
        if not is_zero_row(values):
            extracted_rows.append({"label": label_text, "values": values})

        if label_matches(label_text, ROLLING_IS_STOP_LABEL):
            break

    # -----------------------------------------------------------------------
    # CS-specific: drop month columns where NOI is 0 or None.
    # CubeSmart shows a rolling 12-month view; unused future months have NOI = 0
    # and should be excluded entirely from both the dates list and each row.
    # -----------------------------------------------------------------------
    noi_row = None
    for r in extracted_rows:
        if label_matches(r["label"], ROLLING_IS_STOP_LABEL):
            noi_row = r
            break

    if noi_row is not None:
        keep_cols = []
        for i, val in enumerate(noi_row["values"]):
            if val is None:
                continue
            try:
                if float(val) != 0:
                    keep_cols.append(i)
            except (ValueError, TypeError):
                # Non-numeric NOI cell — unexpected, but keep the column rather than silently drop it
                keep_cols.append(i)

        if len(keep_cols) < len(dates):
            dates = [dates[i] for i in keep_cols]
            for r in extracted_rows:
                r["values"] = [r["values"][i] for i in keep_cols]

    return dates, extracted_rows


# ---------------------------------------------------------------------------
# CORE PROCESSING FUNCTION — ROUTES BY MANAGED_BY
#
# This is the main function that both the CLI script and the webapp call.
# It takes a filepath and property_name, returns (output_bytes, filename, log).
# It does NOT print anything or prompt for input.
# Branches: "Public Storage" → PS extractors, "CubeSmart" → CS extractors,
#           else → EXR extractors.
# Other: uses EXR extraction, skips COA mapping.
# ---------------------------------------------------------------------------

def process_workbook(filepath, property_name, managed_by="Extra"):
    """
    Process a single Excel file and return the output workbook as bytes.

    Args:
        filepath:       Path to the input .xlsx file
        property_name:  Name to use in the Rolling IS tab and output filename

    Returns:
        A dict with:
            "output_bytes":    bytes of the output .xlsx workbook (or None on error)
            "output_filename": suggested filename like 'Chattanooga_datapack.xlsx'
            "log":             list of dicts with keys: sheet, status, message
            "summary":         dict with counts of what was extracted
    """
    filename = os.path.basename(filepath)
    log = []
    summary = {}

    # -- Open the input workbook --
    # read_only=True: faster and uses less memory — we never write back to the source file.
    # data_only=True: returns the last-calculated cell values instead of formula strings.
    #                 Without this, cells with formulas would return None.
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        log.append({"sheet": "", "status": "ERROR", "message": f"Could not open: {e}"})
        return {"output_bytes": None, "output_filename": None, "log": log, "summary": summary}

    # -- Extract from each sheet --

    rolling_is_data = None
    unit_rate_data  = None
    ops_sum_data    = None
    rent_roll_data  = None

    if managed_by == "Public Storage":
        # ---------------------------------------------------------------
        # PUBLIC STORAGE (PS) branch — IS sheet + Rent Roll occupancy count
        # ---------------------------------------------------------------

        # Rolling IS — sheet named "IS"
        ws = wb["IS"] if "IS" in wb.sheetnames else None
        if ws is None:
            log.append({"sheet": "IS", "status": "WARNING",
                         "message": "IS sheet not found"})
        else:
            prop_num = extract_ps_property_number(ws)
            dates, rows = extract_ps_rolling_is(ws)
            if dates is None:
                log.append({"sheet": "IS", "status": "WARNING",
                             "message": "Could not find date header row"})
            elif rows is None:
                log.append({"sheet": "IS", "status": "WARNING",
                             "message": f"Could not find '{ROLLING_IS_START_LABEL}' label"})
            else:
                rolling_is_data = {"prop_num": prop_num, "dates": dates, "rows": rows}
                msg = f"Extracted {len(rows)} line items x {len(dates)} months"
                log.append({"sheet": "IS", "status": "OK", "message": msg})
                summary["rolling_is"] = msg

        # Unit Rate — derived from Rent Roll occupancy count
        ws_rr = wb["Rent Roll"] if "Rent Roll" in wb.sheetnames else None
        if ws_rr is None:
            log.append({"sheet": "Rent Roll", "status": "WARNING",
                         "message": "Rent Roll sheet not found — cannot derive occupancy"})
        else:
            occupied = extract_ps_rent_roll_occupancy(ws_rr)
            if occupied is None:
                log.append({"sheet": "Rent Roll", "status": "WARNING",
                             "message": "Could not count occupied units"})
            else:
                prop_num_rr = prop_num if rolling_is_data else ""
                unit_rate_data = {
                    "prop_num": prop_num_rr,
                    "metrics": {"Units Rented": occupied},
                }
                msg = f"Occupied units: {occupied} (Units Available / Sq Ft not in PS format)"
                log.append({"sheet": "Rent Roll", "status": "OK", "message": msg})
                summary["unit_rate"] = msg

        # Ops Sum and full Rent Roll — not available in PS format
        log.append({"sheet": "Ops Sum",   "status": "SKIP",
                    "message": "Not available in Public Storage format"})
        log.append({"sheet": "Rent Roll detail", "status": "SKIP",
                    "message": "PS Rent Roll does not include rates or move-in dates"})

    elif managed_by == "CubeSmart":
        # ---------------------------------------------------------------
        # CUBESMART (CS) branch — Rolling Details sheet only (first pass).
        # Unit Rate / Ops Sum / Rent Roll not yet mapped for CS.
        # ---------------------------------------------------------------

        # Rolling Details — sheet named exactly "Rolling Details"
        ws = wb[CS_ROLLING_IS_SHEET] if CS_ROLLING_IS_SHEET in wb.sheetnames else None
        if ws is None:
            log.append({"sheet": CS_ROLLING_IS_SHEET, "status": "WARNING",
                         "message": f"{CS_ROLLING_IS_SHEET} sheet not found"})
        else:
            prop_num = extract_cs_property_number(ws)
            dates, rows = extract_cs_rolling_is(ws)
            if dates is None:
                log.append({"sheet": CS_ROLLING_IS_SHEET, "status": "WARNING",
                             "message": "Could not find date header row"})
            elif rows is None:
                log.append({"sheet": CS_ROLLING_IS_SHEET, "status": "WARNING",
                             "message": f"Could not find '{CS_ROLLING_IS_START_LABEL}' label"})
            else:
                rolling_is_data = {"prop_num": prop_num, "dates": dates, "rows": rows}
                msg = f"Extracted {len(rows)} line items x {len(dates)} months"
                log.append({"sheet": CS_ROLLING_IS_SHEET, "status": "OK", "message": msg})
                summary["rolling_is"] = msg

        # Other tabs not yet implemented for CubeSmart
        log.append({"sheet": "Unit Rate", "status": "SKIP",
                    "message": "Not yet implemented for CubeSmart"})
        log.append({"sheet": "Ops Sum",   "status": "SKIP",
                    "message": "Not yet implemented for CubeSmart"})
        log.append({"sheet": "Rent Roll", "status": "SKIP",
                    "message": "Not yet implemented for CubeSmart"})

    else:
        # ---------------------------------------------------------------
        # EXTRA SPACE (EXR) branch (also used by "Other" until that format
        # gets its own extraction logic)
        # ---------------------------------------------------------------

        # Rolling IS
        sheet_name, ws = find_sheet_by_prefix(wb, SHEET_PREFIXES["rolling_is"])
        if ws is None:
            log.append({"sheet": "Rolling IS", "status": "WARNING",
                         "message": "Sheet not found"})
        else:
            prop_num = extract_property_number(sheet_name, SHEET_PREFIXES["rolling_is"])
            dates, rows = extract_rolling_is(ws)
            if dates is None:
                log.append({"sheet": sheet_name, "status": "WARNING",
                             "message": "Could not find date header row"})
            elif rows is None:
                log.append({"sheet": sheet_name, "status": "WARNING",
                             "message": f"Could not find '{EXR_ROLLING_IS_START_LABEL}' label"})
            else:
                rolling_is_data = {"prop_num": prop_num, "dates": dates, "rows": rows}
                msg = f"Extracted {len(rows)} line items x {len(dates)} months"
                log.append({"sheet": sheet_name, "status": "OK", "message": msg})
                summary["rolling_is"] = msg

        # Unit Rate
        sheet_name, ws = find_sheet_by_prefix(wb, SHEET_PREFIXES["unit_rate"])
        if ws is None:
            log.append({"sheet": "Unit Rate", "status": "WARNING",
                         "message": "Sheet not found"})
        else:
            prop_num = extract_property_number(sheet_name, SHEET_PREFIXES["unit_rate"])
            metrics = extract_unit_rate(ws)
            if not metrics:
                log.append({"sheet": sheet_name, "status": "WARNING",
                             "message": "No matching metrics found"})
            else:
                missing = set(UNIT_RATE_LABELS) - set(metrics.keys())
                if missing:
                    log.append({"sheet": sheet_name, "status": "WARNING",
                                 "message": f"Missing: {', '.join(missing)}"})
                unit_rate_data = {"prop_num": prop_num, "metrics": metrics}
                msg = f"Extracted {len(metrics)} metrics"
                log.append({"sheet": sheet_name, "status": "OK", "message": msg})
                summary["unit_rate"] = msg

        # Ops Sum
        sheet_name, ws = find_sheet_by_prefix(wb, SHEET_PREFIXES["ops_sum"])
        if ws is None:
            log.append({"sheet": "Ops Sum", "status": "WARNING",
                         "message": "Sheet not found"})
        else:
            prop_num = extract_property_number(sheet_name, SHEET_PREFIXES["ops_sum"])
            dates, rows = extract_ops_sum(ws)
            if dates is None:
                log.append({"sheet": sheet_name, "status": "WARNING",
                             "message": "Could not find date header row"})
            elif rows is None:
                log.append({"sheet": sheet_name, "status": "WARNING",
                             "message": "Could not find label column"})
            else:
                missing = set(OPS_SUM_LABELS) - {r["label"] for r in rows}
                if missing:
                    log.append({"sheet": sheet_name, "status": "WARNING",
                                 "message": f"Missing: {', '.join(missing)}"})
                ops_sum_data = {"prop_num": prop_num, "dates": dates, "rows": rows}
                msg = f"Extracted {len(rows)} metrics x {len(dates)} months"
                log.append({"sheet": sheet_name, "status": "OK", "message": msg})
                summary["ops_sum"] = msg

        # Rent Roll
        sheet_name, ws = find_sheet_by_prefix(wb, SHEET_PREFIXES["rent_roll"])
        if ws is None:
            log.append({"sheet": "Rent Roll", "status": "WARNING",
                         "message": "Sheet not found"})
        else:
            prop_num = extract_property_number(sheet_name, SHEET_PREFIXES["rent_roll"])
            headers, data_rows = extract_rent_roll(ws)
            if headers is None:
                log.append({"sheet": sheet_name, "status": "WARNING",
                             "message": "Could not find header row"})
            elif not data_rows:
                log.append({"sheet": sheet_name, "status": "WARNING",
                             "message": "No tenant/unit rows found"})
            else:
                # Run ECRI / mark-to-market analytics — appends PSF and delta columns
                enh_headers, enh_rows, rr_summary = calculate_rent_roll_analytics(
                    headers, data_rows
                )
                rent_roll_data = {
                    "prop_num":  prop_num,
                    "headers":   enh_headers,
                    "data_rows": enh_rows,
                    "summary":   rr_summary,
                }
                msg = (f"Extracted {len(enh_rows)} tenants x {len(enh_headers)} columns "
                       f"({rr_summary['below_street_count']} below street rate)")
                log.append({"sheet": sheet_name, "status": "OK", "message": msg})
                summary["rent_roll"] = msg

    wb.close()

    # -- Run COA mapping on extracted Rolling IS accounts (if mapper is available) --
    coa_lookup = {}   # {label: result_dict} — populated below if mapper loaded
    approved_file, alias_file = COA_FILES.get(managed_by, (None, None))

    if rolling_is_data and managed_by == "Other":
        log.append({"sheet": "COA Mapper", "status": "SKIP",
                    "message": "Managed By = Other — COA mapping is manual for this property"})
    elif rolling_is_data and approved_file is None:
        log.append({"sheet": "COA Mapper", "status": "SKIP",
                    "message": f"No COA mapping file configured for '{managed_by}'"})
    elif rolling_is_data and _COA_MAPPER_AVAILABLE:
        approved_path = os.path.join(_HERE_CORE, approved_file)
        alias_path    = os.path.join(_HERE_CORE, alias_file)
        mapper     = COAMapper(approved_file=approved_path, alias_file=alias_path)
        coa_lookup = mapper.map_unique_from_rows(rolling_is_data["rows"])
        auto_ok    = sum(1 for r in coa_lookup.values() if not r["review_required"])
        need_rev   = sum(1 for r in coa_lookup.values() if  r["review_required"])
        msg = f"COA mapping: {auto_ok} auto-accepted, {need_rev} flagged for review"
        log.append({"sheet": "COA Mapper", "status": "OK", "message": msg})
        summary["coa_mapping"] = msg
    elif rolling_is_data and not _COA_MAPPER_AVAILABLE:
        log.append({"sheet": "COA Mapper", "status": "SKIP",
                    "message": "coa_mapper.py not found — COA Mapping tabs skipped"})

    # -- Build the output workbook --
    # openpyxl always creates a new Workbook with one blank "Sheet" tab.
    # We remove it immediately so the output only contains our named tabs.
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    if rolling_is_data:
        write_rolling_is_tab(out_wb, filename, rolling_is_data["prop_num"],
                             rolling_is_data["dates"], rolling_is_data["rows"],
                             property_name=property_name)
        if coa_lookup:
            write_rolling_is_mapped_tab(out_wb, filename, rolling_is_data["prop_num"],
                                        rolling_is_data["dates"], rolling_is_data["rows"],
                                        property_name, coa_lookup)

    if unit_rate_data:
        write_unit_rate_tab(out_wb, filename, unit_rate_data["prop_num"],
                            unit_rate_data["metrics"])

    if ops_sum_data:
        write_ops_sum_tab(out_wb, filename, ops_sum_data["prop_num"],
                          ops_sum_data["dates"], ops_sum_data["rows"])

    if rent_roll_data:
        write_rent_roll_tab(out_wb, filename, rent_roll_data["prop_num"],
                            rent_roll_data["headers"], rent_roll_data["data_rows"],
                            summary=rent_roll_data.get("summary"))

    if coa_lookup:
        write_coa_mapping_tab(out_wb, list(coa_lookup.values()))

    # Convert log dicts to lists for the log tab (writer expects plain lists, not dicts)
    log_list = [[datetime.now().isoformat(), e["sheet"], e["status"], e["message"]]
                for e in log]
    write_log_tab(out_wb, log_list)

    # -- Save to bytes so the caller can write to disk or stream to a browser --
    # BytesIO is an in-memory file buffer. We save the workbook into it,
    # then read back the raw bytes. The CLI writes those bytes to a .xlsx file;
    # the webapp streams them as a download — neither needs a temp file on disk.
    from io import BytesIO
    buffer = BytesIO()
    out_wb.save(buffer)
    output_bytes = buffer.getvalue()

    safe_name = make_safe_filename(property_name)
    output_filename = f"{safe_name}_datapack.xlsx"

    return {
        # Core output — used by both the CLI script and the webapp
        "output_bytes":    output_bytes,    # the .xlsx file as raw bytes
        "output_filename": output_filename, # suggested filename (e.g. "Chattanooga_datapack.xlsx")
        "log":             log,             # list of {sheet, status, message} dicts
        "summary":         summary,         # short human-readable extraction counts
        # Extended fields — used by db_writer.py to populate the db_ready/ CSVs
        "rolling_is_data": rolling_is_data,
        "unit_rate_data":  unit_rate_data,
        "ops_sum_data":    ops_sum_data,
        "rent_roll_data":  rent_roll_data,
        "coa_lookup":      coa_lookup,      # {source_label: mapping_result} from COA mapper
        "managed_by":      managed_by,      # echoed back so callers can branch on it
    }


# ---------------------------------------------------------------------------
# EXCEL OUTPUT WRITERS — SHARED (all formats)
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True)
NUMBER_FORMAT = '#,##0.00'
INTEGER_FORMAT = '#,##0'
DATE_FORMAT = 'm/d/yyyy'

# Cell fill colors for COA confidence visualization in the mapping tabs.
# These match Excel's built-in Good / Neutral / Bad conditional format palette.
FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")  # auto-accepted (>= 0.85)
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")  # needs review, suggestion exists
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")  # no mapping found

# ---------------------------------------------------------------------------
# MANAGED BY — ALL FORMATS (Extra, Public Storage, CubeSmart, Other)
# ---------------------------------------------------------------------------

MANAGED_BY_OPTIONS = [
    "Extra",          # Extra Space Storage — EXR format
    "Public Storage", # Public Storage format
    "CubeSmart",      # CubeSmart format
    "Other",          # Smaller / unknown managers — COA mapping is manual
]
DEFAULT_MANAGED_BY = "Extra"

# Maps each managed_by option to its approved and alias mapping CSV files.
# None means skip COA mapping for that manager (manual review required).
_HERE_CORE = os.path.dirname(os.path.abspath(__file__))
COA_FILES = {
    "Extra":          ("approved_mappings_exr.csv", "alias_mappings_exr.csv"),
    "Public Storage": ("approved_mappings_ps.csv",  "alias_mappings_ps.csv"),
    "CubeSmart":      ("approved_mappings_cs.csv",  "alias_mappings_cs.csv"),
    "Other":          (None, None),
}


# ---------------------------------------------------------------------------
# COA MAPPING OUTPUT WRITERS — SHARED (all formats)
# ---------------------------------------------------------------------------

def write_coa_mapping_tab(out_wb, mapping_results):
    """
    Write the COA Mapping review tab — one row per unique source account.

    Sorted: Income accounts first, then Expense, then EXR_Rollup.
    Color-coded by confidence so high-priority reviews are visible at a glance.

    Columns:
      A  Source Account    E  Confidence
      B  Suggested COA     F  Match Method
      C  Suggested COA 2   G  Review Required
      D  Account Type      H  Notes
    """
    ws = out_wb.create_sheet(title="COA Mapping")

    ws["A1"] = "COA Mapping Review"
    ws["A1"].font = HEADER_FONT

    # Sort Income first, then Expense, then Rollup rows, then anything unrecognised.
    # Within each group, sort alphabetically by source label.
    # The lambda returns a tuple: Python sorts tuples element by element,
    # so account type is the primary sort key and label is the tiebreaker.
    TYPE_ORDER = {"Income": 0, "Expense": 1, "EXR_Rollup": 2}
    sorted_results = sorted(
        mapping_results,
        key=lambda r: (TYPE_ORDER.get(r.get("account_type", ""), 3),
                       r.get("source_label", ""))
    )

    header_row = 3
    col_headers = ["Source Account", "Suggested COA", "Suggested COA 2",
                   "Account Type", "Confidence", "Match Method",
                   "Review Required", "Notes"]
    for col_idx, header in enumerate(col_headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header).font = HEADER_FONT

    for row_idx, result in enumerate(sorted_results, start=4):
        confidence = result.get("confidence", 0.0)
        review     = result.get("review_required", True)

        # Green  = auto-accepted: high confidence, no review needed
        # Yellow = suggestion exists but needs a human to confirm
        # Red    = no match found at all
        if not review and confidence >= _COA_AUTO_ACCEPT:
            fill = FILL_GREEN
        elif result.get("coa"):
            fill = FILL_YELLOW
        else:
            fill = FILL_RED

        values = [
            result.get("source_label", ""),
            result.get("coa", ""),
            result.get("coa2", ""),
            result.get("account_type", ""),
            confidence,
            result.get("match_method", ""),
            "YES" if review else "NO",
            result.get("notes", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            if col_idx == 5:
                cell.number_format = "0%"

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 17
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 65
    # freeze_panes locks rows 1-3 and column A so the header stays visible while scrolling
    ws.freeze_panes = "A4"
    return ws


def write_rolling_is_mapped_tab(out_wb, source_file, prop_num, dates, rows,
                                 property_name, coa_lookup):
    """
    Write Rolling IS data with COA and COA 2 columns added (after line_item).

    Same structure as the Rolling IS tab but with COA context columns so this
    tab can be used directly as a model input without manual lookups.

    Columns: Property Name, line_item, COA, COA 2, Month, Year, Period, Amount
    The COA cell is color-coded by confidence (green/yellow/red).
    """
    ws = out_wb.create_sheet(title="Rolling IS Mapped")

    ws["A1"] = "Source File"
    ws["B1"] = source_file
    ws["A2"] = "Property Number"
    ws["B2"] = prop_num
    ws["A1"].font = HEADER_FONT
    ws["A2"].font = HEADER_FONT

    header_row = 4
    col_headers = ["Property Name", "line_item", "COA", "COA 2",
                   "Month", "Year", "Period", "Amount"]
    for col_idx, header in enumerate(col_headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header).font = HEADER_FONT

    data_row = header_row + 1
    for row_data in rows:
        label      = row_data["label"]
        mapping    = coa_lookup.get(label, {})
        coa_val    = mapping.get("coa", "")
        coa2_val   = mapping.get("coa2", "")
        confidence = mapping.get("confidence", 0.0)
        review     = mapping.get("review_required", True)

        if coa_val and not review and confidence >= _COA_AUTO_ACCEPT:
            coa_fill = FILL_GREEN
        elif coa_val:
            coa_fill = FILL_YELLOW
        else:
            coa_fill = FILL_RED

        for i, date_str in enumerate(dates):
            month, year, period_date = parse_date_string(date_str)
            amount = row_data["values"][i] if i < len(row_data["values"]) else 0

            ws.cell(row=data_row, column=1, value=property_name if label else "")
            ws.cell(row=data_row, column=2, value=label)

            coa_cell = ws.cell(row=data_row, column=3, value=coa_val)
            coa_cell.fill = coa_fill

            ws.cell(row=data_row, column=4, value=coa2_val)
            ws.cell(row=data_row, column=5, value=month)
            ws.cell(row=data_row, column=6, value=year)

            period_cell = ws.cell(row=data_row, column=7, value=period_date)
            period_cell.number_format = DATE_FORMAT

            amount_cell = ws.cell(row=data_row, column=8,
                                  value=amount if amount is not None else 0)
            amount_cell.number_format = NUMBER_FORMAT

            data_row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14
    ws.freeze_panes = "A5"
    return ws


def write_rolling_is_tab(out_wb, source_file, prop_num, dates, rows, property_name=""):
    """
    Write Rolling IS in long (database-friendly) format.

    Each source row has one month per column. We "unpivot" that into one output
    row per account × month combination, which makes the data easy to filter,
    pivot, and load into a database or the proforma Data Drop sheet.

    Example: 26 accounts × 12 months = 312 output rows.
    """
    ws = out_wb.create_sheet(title="Rolling IS")

    ws["A1"] = "Source File"
    ws["B1"] = source_file
    ws["A2"] = "Property Number"
    ws["B2"] = prop_num
    ws["A1"].font = HEADER_FONT
    ws["A2"].font = HEADER_FONT

    header_row = 4
    headers = ["Property Name", "line_item", "Month", "Year", "Period", "Amount"]
    for col_idx, header in enumerate(headers):
        ws.cell(row=header_row, column=col_idx + 1, value=header).font = HEADER_FONT

    data_row = header_row + 1
    for row_data in rows:
        # For each account row, loop through every month and write one output row
        for i, date_str in enumerate(dates):
            month, year, period_date = parse_date_string(date_str)
            amount = row_data["values"][i] if i < len(row_data["values"]) else 0

            line_item = row_data["label"]
            # Only populate Property Name when the row has a label (avoids blank rows)
            ws.cell(row=data_row, column=1, value=property_name if line_item else "")
            ws.cell(row=data_row, column=2, value=line_item)
            ws.cell(row=data_row, column=3, value=month)
            ws.cell(row=data_row, column=4, value=year)

            period_cell = ws.cell(row=data_row, column=5, value=period_date)
            period_cell.number_format = DATE_FORMAT

            amount_cell = ws.cell(row=data_row, column=6, value=amount if amount is not None else 0)
            amount_cell.number_format = NUMBER_FORMAT

            data_row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    return ws


def calculate_rent_roll_analytics(headers, data_rows):
    """
    Adds ECRI / mark-to-market calculated columns to the extracted rent roll.

    For each tenant row, calculates:
      Rent Rate PSF    — in-place rent divided by square footage
      Street Rate PSF  — asking rent divided by square footage
      Delta to Street  — Street Rate minus Rent Rate (positive = below market)
      Delta PSF        — Street Rate PSF minus Rent Rate PSF
      Below Street     — 1 if Rent Rate < Street Rate, else 0

    Also returns a summary dict with portfolio-level metrics.

    Returns:
        enhanced_headers  — original headers + 5 new column names
        enhanced_rows     — original rows + 5 new values per row
        summary           — dict with aggregate metrics
    """

    # Find column positions by name so this is not sensitive to column order
    def col_idx(name):
        return headers.index(name) if name in headers else None

    i_rent   = col_idx("Rent Rate")
    i_street = col_idx("Street Rate")
    i_sqft   = col_idx("Sq Ft")
    i_status = col_idx("Status")

    new_cols = [
        "Rent Rate PSF",
        "Street Rate PSF",
        "Delta to Street Rate",
        "Delta PSF",
        "Below Street Rate",
    ]
    enhanced_headers = list(headers) + new_cols

    enhanced_rows   = []
    current_count   = 0    # tenants whose Status = "Current"
    rent_psf_vals   = []   # for avg Rent PSF
    street_psf_vals = []   # for avg Street PSF
    positive_deltas = []   # delta > 0 for each below-street tenant

    for row in data_rows:
        # Safely pull numeric values; treat None / non-numeric as None
        def safe_num(idx):
            if idx is None or idx >= len(row):
                return None
            try:
                return float(row[idx]) if row[idx] is not None else None
            except (ValueError, TypeError):
                return None

        rent   = safe_num(i_rent)
        street = safe_num(i_street)
        sqft   = safe_num(i_sqft)

        # Count tenants with Status = "Current" for the Occupied Tenants summary line
        if i_status is not None and i_status < len(row):
            if str(row[i_status] or "").strip().lower() == "current":
                current_count += 1

        # PSF — only valid when sq_ft is a positive number
        rent_psf   = (rent   / sqft) if (rent   is not None and sqft and sqft > 0) else None
        street_psf = (street / sqft) if (street is not None and sqft and sqft > 0) else None

        # Delta — dollar gap between asking rent and in-place rent
        delta     = (street - rent)         if (rent is not None and street is not None) else None
        delta_psf = (street_psf - rent_psf) if (rent_psf is not None and street_psf is not None) else None

        # Flag — 1 if the tenant is paying below the current street rate
        below_flag = 1 if (rent is not None and street is not None and rent < street) else 0

        # Accumulate for summary
        if rent_psf   is not None: rent_psf_vals.append(rent_psf)
        if street_psf is not None: street_psf_vals.append(street_psf)
        if delta is not None and delta > 0:
            positive_deltas.append(delta)

        enhanced_rows.append(list(row) + [rent_psf, street_psf, delta, delta_psf, below_flag])

    # Portfolio-level summary
    occupied_count       = current_count
    below_street_count   = len(positive_deltas)
    pct_below_street     = (below_street_count / occupied_count) if occupied_count > 0 else None
    total_positive_delta = sum(positive_deltas) if positive_deltas else 0
    avg_positive_delta   = (total_positive_delta / below_street_count) if below_street_count > 0 else None
    avg_rent_psf         = (sum(rent_psf_vals)   / len(rent_psf_vals))   if rent_psf_vals   else None
    avg_street_psf       = (sum(street_psf_vals) / len(street_psf_vals)) if street_psf_vals else None

    summary = {
        "occupied_count":       occupied_count,
        "below_street_count":   below_street_count,
        "pct_below_street":     pct_below_street,
        "total_positive_delta": total_positive_delta,
        "avg_positive_delta":   avg_positive_delta,
        "avg_rent_psf":         avg_rent_psf,
        "avg_street_psf":       avg_street_psf,
    }

    return enhanced_headers, enhanced_rows, summary


def write_unit_rate_tab(out_wb, source_file, prop_num, metrics):
    """Write Unit Rate as a simple metric/value table."""
    ws = out_wb.create_sheet(title="Unit Rate")

    ws["A1"] = "Source File"
    ws["B1"] = source_file
    ws["A2"] = "Property Number"
    ws["B2"] = prop_num
    ws["A1"].font = HEADER_FONT
    ws["A2"].font = HEADER_FONT

    ws.cell(row=4, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=4, column=2, value="Value").font = HEADER_FONT

    row_num = 5
    for label in UNIT_RATE_LABELS:
        if label in metrics:
            ws.cell(row=row_num, column=1, value=label)
            ws.cell(row=row_num, column=2, value=metrics[label]).number_format = INTEGER_FORMAT
            row_num += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    return ws


def write_ops_sum_tab(out_wb, source_file, prop_num, dates, rows):
    """Write Ops Sum in long format."""
    ws = out_wb.create_sheet(title="Ops Sum")

    ws["A1"] = "Source File"
    ws["B1"] = source_file
    ws["A2"] = "Property Number"
    ws["B2"] = prop_num
    ws["A1"].font = HEADER_FONT
    ws["A2"].font = HEADER_FONT

    header_row = 4
    headers = ["metric", "Month", "Year", "Period", "Value"]
    for col_idx, header in enumerate(headers):
        ws.cell(row=header_row, column=col_idx + 1, value=header).font = HEADER_FONT

    data_row = header_row + 1
    for row_data in rows:
        for i, date_str in enumerate(dates):
            month, year, period_date = parse_date_string(date_str)
            value = row_data["values"][i] if i < len(row_data["values"]) else 0

            ws.cell(row=data_row, column=1, value=row_data["label"])
            ws.cell(row=data_row, column=2, value=month)
            ws.cell(row=data_row, column=3, value=year)
            ws.cell(row=data_row, column=4, value=period_date).number_format = DATE_FORMAT
            ws.cell(row=data_row, column=5, value=value if value is not None else 0).number_format = INTEGER_FORMAT

            data_row += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    return ws


def write_rent_roll_tab(out_wb, source_file, prop_num, headers, data_rows,
                        summary=None):
    """
    Write Rent Roll as a flat table.

    When a summary dict is provided (from calculate_rent_roll_analytics),
    a compact summary block is written above the data table so ECRI / mark-to-market
    metrics are visible as soon as the tab opens.

    Layout with summary:
      Rows 1-2:  source file / property metadata
      Rows 4-11: summary block (8 rows: label + 7 metrics)
      Row 13:    table column headers
      Row 14+:   tenant data rows

    Without summary (e.g. PS occupancy-only): original compact layout.
    """
    ws = out_wb.create_sheet(title="Rent Roll")

    # -- Metadata rows (always present) --
    ws["A1"] = "Source File"
    ws["B1"] = source_file
    ws["A2"] = "Property Number"
    ws["B2"] = prop_num
    ws["A1"].font = HEADER_FONT
    ws["A2"].font = HEADER_FONT

    if summary:
        # -- Summary block --
        ws["A4"] = "Rent Roll Summary"
        ws["A4"].font = HEADER_FONT

        summary_rows = [
            ("Occupied Tenants",                  summary.get("occupied_count"),       "integer"),
            ("Below Street Rate",                 summary.get("below_street_count"),   "integer"),
            ("% Below Street",                    summary.get("pct_below_street"),     "percent"),
            ("Total Positive Delta to Street",    summary.get("total_positive_delta"), "money"),
            ("Avg Delta per Below-Street Tenant", summary.get("avg_positive_delta"),   "money"),
            ("Avg Rent PSF",                      summary.get("avg_rent_psf"),         "money"),
            ("Avg Street PSF",                    summary.get("avg_street_psf"),       "money"),
        ]

        for i, (label, value, fmt) in enumerate(summary_rows):
            row_num = 5 + i
            ws.cell(row=row_num, column=1, value=label).font = HEADER_FONT
            cell = ws.cell(row=row_num, column=2, value=value)
            if fmt == "integer":
                cell.number_format = INTEGER_FORMAT
            elif fmt == "percent":
                cell.number_format = "0.0%"   # stored as decimal, e.g. 0.35 shows as 35.0%
            else:
                cell.number_format = NUMBER_FORMAT

        header_row = 13   # table headers start here (7 summary rows + blank gap)
    else:
        header_row = 4    # original compact layout (no summary)

    # -- Column headers --
    for col_idx, header in enumerate(headers):
        ws.cell(row=header_row, column=col_idx + 1, value=header).font = HEADER_FONT

    # Which columns get which number format
    date_columns    = {"Move-In Date", "Paid-Thru Date"}
    money_columns   = {"Rent Rate", "Street Rate",
                       "Rent Rate PSF", "Street Rate PSF",
                       "Delta to Street Rate", "Delta PSF"}
    integer_columns = {"Sq Ft", "Below Street Rate"}

    # -- Data rows --
    for row_idx, row_values in enumerate(data_rows):
        data_row = header_row + 1 + row_idx
        for col_idx, val in enumerate(row_values):
            cell = ws.cell(row=data_row, column=col_idx + 1, value=val)
            col_name = headers[col_idx] if col_idx < len(headers) else ""
            if col_name in date_columns and isinstance(val, datetime):
                cell.number_format = DATE_FORMAT
            elif col_name in money_columns and val is not None:
                cell.number_format = NUMBER_FORMAT
            elif col_name in integer_columns and val is not None:
                cell.number_format = INTEGER_FORMAT

    # -- Column widths --
    col_widths = {
        "Tenant Account":        16,
        "Unit #":                10,
        "Move-In Date":          13,
        "Rent Rate":             12,
        "Street Rate":           12,
        "Paid-Thru Date":        15,
        "Status":                10,
        "Size":                   8,
        "Type":                   8,
        "Sq Ft":                  8,
        "Rent Rate PSF":         14,
        "Street Rate PSF":       14,
        "Delta to Street Rate":  18,
        "Delta PSF":             12,
        "Below Street Rate":     16,
    }
    for col_idx, header in enumerate(headers):
        # openpyxl uses letters for columns; for more than 26 cols use get_column_letter
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_idx + 1)
        ws.column_dimensions[col_letter].width = col_widths.get(header, 12)

    return ws


def write_log_tab(out_wb, log_entries):
    """Write processing log tab."""
    ws = out_wb.create_sheet(title="Processing Log")

    headers = ["Timestamp", "Sheet", "Status", "Message"]
    for col_idx, header in enumerate(headers):
        ws.cell(row=1, column=col_idx + 1, value=header).font = HEADER_FONT

    for row_idx, entry in enumerate(log_entries):
        for col_idx, val in enumerate(entry):
            ws.cell(row=row_idx + 2, column=col_idx + 1, value=val)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 55
    return ws
