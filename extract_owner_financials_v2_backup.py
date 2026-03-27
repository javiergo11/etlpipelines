"""
extract_owner_financials.py
===========================
Extracts historical property data from Extra Space Storage owner financial
Excel files and writes one clean .xlsx output workbook per input file.
 
HOW IT WORKS:
  1. Looks in the "input" folder for any .xlsx files
  2. For each file, prompts you for a property name (used in the output filename)
  3. Finds sheets by prefix (Rolling IS, Unit Rate, Ops Sum)
  4. Extracts data using label-based search (not hardcoded cell addresses)
  5. Writes one output workbook per input file to the "output" folder
     with tabs: Rolling IS, Unit Rate, Ops Sum, Processing Log
 
USAGE:
  - Drop .xlsx files into the "input" folder
  - Run:  python extract_owner_financials.py
  - Follow the prompts to name each output file
  - Check the "output" folder for results
"""
 
import os
import sys
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
 
 
# ---------------------------------------------------------------------------
# CONFIGURATION  (edit these if your files use different label text)
# ---------------------------------------------------------------------------
 
# Sheet prefix -> we match sheets whose name starts with this text
SHEET_PREFIXES = {
    "rolling_is": "Rolling IS",
    "unit_rate":  "Unit Rate",
    "ops_sum":    "Ops Sum",
    "rent_roll":  "Rent Roll",
}
 
# Rolling IS: labels that mark the start and end of the section we want
ROLLING_IS_START_LABEL = "Rental Income"  # first line item to capture
ROLLING_IS_STOP_LABEL  = "Net Operating Income"  # last line item to capture
 
# Unit Rate: the summary metrics we want to pull
UNIT_RATE_LABELS = [
    "Units Available",
    "Units Rented",
    "Sq Ft Available",
    "Sq Ft Rented",
]
 
# Ops Sum: the rental activity rows we want
OPS_SUM_LABELS = [
    "Rentals During Month",
    "Walk In Rentals",
    "NSC rentals",       # note: lowercase 'r' in the actual file
    "Web Rentals",
    "Vacates During Month",
    "Net Rentals",
]
 
# Rent Roll: the column headers we expect to find in the header row.
# The script searches for a row containing these labels (in any column)
# rather than assuming they are always in row 5 or columns B:J.
RENT_ROLL_HEADERS = [
    "Tenant Account",
    "Unit #",
    "Move-In Date",
    "Rent Rate",
    "Street Rate",
    "Paid-Thru Date",
    "Status",
    "Size",
    "Type",
]
 
 
# ---------------------------------------------------------------------------
# HELPER FUNCTIONS
# ---------------------------------------------------------------------------
 
def setup_folders():
    """Create 'input' and 'output' folders if they don't already exist."""
    for folder in ("input", "output"):
        if not os.path.exists(folder):
            os.makedirs(folder)
            print(f"  Created folder: {folder}/")
 
 
def find_xlsx_files(folder):
    """Return a list of .xlsx file paths inside the given folder."""
    files = []
    for name in sorted(os.listdir(folder)):
        if name.lower().endswith(".xlsx") and not name.startswith("~$"):
            files.append(os.path.join(folder, name))
    return files
 
 
def find_sheet_by_prefix(workbook, prefix):
    """
    Find the first sheet whose name starts with the given prefix.
    Returns (sheet_name, worksheet) or (None, None) if not found.
    """
    for name in workbook.sheetnames:
        if name.startswith(prefix):
            return name, workbook[name]
    return None, None
 
 
def extract_property_number(sheet_name, prefix):
    """
    Pull the property number from a sheet name like 'Rolling IS 7214'.
    Returns the numeric suffix as a string, or 'UNKNOWN'.
    """
    remainder = sheet_name.replace(prefix, "").strip()
    digits = re.sub(r"[^0-9]", "", remainder)
    return digits if digits else "UNKNOWN"
 
 
def clean_label(text):
    """
    Normalize a cell label for comparison:
      - strip whitespace
      - lowercase
    """
    if text is None:
        return ""
    return str(text).strip().lower()
 
 
def label_matches(cell_text, target):
    """
    Check if a cell's text starts with the target label.
    For example, 'Rental Income (4000)' matches 'Rental Income'.
    Comparison is case-insensitive.
    """
    return clean_label(cell_text).startswith(clean_label(target))
 
 
def is_date_value(value):
    """
    Check if a cell value looks like a date (datetime object or a string
    matching 'Mon YYYY' like 'Feb 2025').  Used to separate real date
    columns from trailing text columns like 'YoY Delta'.
    """
    if isinstance(value, datetime):
        return True
    if isinstance(value, str) and re.match(r"[A-Z][a-z]{2}\s+\d{4}", value):
        return True
    return False
 
 
def format_date(value):
    """
    Convert a date cell to a string like 'Feb 2025'.
    Handles datetime objects and strings.
    """
    if isinstance(value, datetime):
        return value.strftime("%b %Y")
    if value is None:
        return ""
    return str(value).strip()
 
 
def is_zero_row(values):
    """
    Return True if every numeric value in the list is 0 (or None).
    Used to filter out empty line items from Rolling IS.
    """
    for v in values:
        if v is None:
            continue
        try:
            if float(v) != 0:
                return False
        except (ValueError, TypeError):
            continue
    return True
 
 
def make_safe_filename(text):
    """
    Remove characters that are invalid in Windows filenames.
    Also collapse multiple spaces/underscores and trim the result.
    """
    # Remove characters not allowed in Windows filenames:  \\ / : * ? " < > |
    safe = re.sub(r'[\\/:*?"<>|]', "", text)
    # Replace spaces with underscores for cleaner filenames
    safe = safe.replace(" ", "_")
    # Collapse multiple underscores
    safe = re.sub(r"_+", "_", safe)
    # Remove leading/trailing underscores
    safe = safe.strip("_")
    return safe
 
 
def guess_property_name(filename):
    """
    Try to guess a reasonable default property name from the input filename.
 
    Examples:
      'Feb. 2026 Owner Financials - EXR Chattanooga.xlsx'  -> 'EXR Chattanooga'
      'Feb__2026_Owner_Financials_-_EXR_Chattanooga.xlsx'  -> 'EXR Chattanooga'
 
    Strategy: normalize underscores to spaces, then split on ' - ' and take
    the part after the last separator.
    """
    # Remove extension
    name = os.path.splitext(filename)[0]
 
    # Normalize underscores to spaces (common in downloaded filenames)
    name = name.replace("_", " ")
 
    # Collapse multiple spaces
    name = re.sub(r" +", " ", name)
 
    # If there's a ' - ' separator, take the part after the last one
    if " - " in name:
        name = name.rsplit(" - ", 1)[1]
    elif " – " in name:  # en-dash variant
        name = name.rsplit(" – ", 1)[1]
 
    return name.strip()
 
 
def prompt_for_property_name(filename):
    """
    Ask the user to enter a property name for the output file.
    Shows a default guess based on the input filename.
    If the user just presses Enter, uses the default.
    """
    default = guess_property_name(filename)
    print(f"\n  Enter property name for output file [default: {default}]: ", end="")
 
    try:
        user_input = input().strip()
    except EOFError:
        user_input = ""
 
    # Use the default if the user didn't type anything
    chosen = user_input if user_input else default
    return chosen
 
 
# ---------------------------------------------------------------------------
# EXTRACTION: Rolling IS
# ---------------------------------------------------------------------------
 
def extract_rolling_is(ws):
    """
    Extract the income statement section from Rental Income to Net Operating
    Income.  Returns (dates, rows) where:
      - dates is a list of date strings (one per month column)
      - rows is a list of dicts: {label, values: [v1, v2, ...]}
 
    Strategy:
      1. Find the date row by scanning for a row where multiple cells are
         dates (datetime objects or strings like 'Feb 2025').
      2. Find the label column by scanning for 'Rental Income'.
      3. Collect rows from Rental Income through Net Operating Income.
      4. Drop rows where all monthly values are zero.
    """
 
    # -- Step 1: Read all rows into a list so we can scan freely --
    all_rows = []
    for row in ws.iter_rows(values_only=False):
        cells = []
        for c in row:
            cells.append(c.value)
        all_rows.append(cells)
 
    if not all_rows:
        return None, None
 
    # -- Step 2: Find the date header row --
    # We look for a row where at least 5 cells are datetime objects or
    # match a month-year pattern like "Feb 2025".
    date_row_idx = None
    date_start_col = None
    date_end_col = None
    dates = []
 
    for idx, row in enumerate(all_rows):
        date_count = 0
        first_date_col = None
        for col_idx, val in enumerate(row):
            if is_date_value(val):
                date_count += 1
                if first_date_col is None:
                    first_date_col = col_idx
        if date_count >= 5:
            date_row_idx = idx
            date_start_col = first_date_col
            # Collect only actual date cells (skip trailing text like 'YoY Delta')
            for col_idx in range(first_date_col, len(row)):
                val = row[col_idx]
                if is_date_value(val):
                    dates.append(format_date(val))
                    date_end_col = col_idx
                elif val is None:
                    break  # stop at first empty cell
                else:
                    break  # stop at non-date text (e.g. 'YoY Delta')
            break
 
    if date_row_idx is None:
        return None, None
 
    # -- Step 3: Find the label column --
    # Scan column by column in the rows near the date row looking for
    # 'Rental Income'.  Usually it's in column C (index 2).
    label_col = None
    start_row_idx = None
 
    for idx in range(date_row_idx + 1, min(date_row_idx + 30, len(all_rows))):
        row = all_rows[idx]
        for col_idx in range(min(5, len(row))):  # labels are in first few cols
            if label_matches(row[col_idx], ROLLING_IS_START_LABEL):
                label_col = col_idx
                start_row_idx = idx
                break
        if label_col is not None:
            break
 
    if label_col is None:
        return dates, None
 
    # -- Step 4: Collect rows from start label to stop label --
    extracted_rows = []
    num_date_cols = len(dates)
 
    for idx in range(start_row_idx, len(all_rows)):
        row = all_rows[idx]
        label_val = row[label_col] if label_col < len(row) else None
 
        if label_val is None or str(label_val).strip() == "":
            continue  # skip blank rows
 
        label_text = str(label_val).strip()
 
        # Grab the monthly values aligned with the date columns
        values = []
        for d in range(num_date_cols):
            col = date_start_col + d
            val = row[col] if col < len(row) else None
            values.append(val)
 
        # Skip rows where all values are zero
        if not is_zero_row(values):
            extracted_rows.append({
                "label": label_text,
                "values": values,
            })
 
        # If this row is NOI, we're done
        if label_matches(label_text, ROLLING_IS_STOP_LABEL):
            break
 
    return dates, extracted_rows
 
 
# ---------------------------------------------------------------------------
# EXTRACTION: Unit Rate
# ---------------------------------------------------------------------------
 
def extract_unit_rate(ws):
    """
    Extract summary metrics (Units Available, Units Rented, etc.) from the
    Unit Rate sheet.  Returns a dict like {'Units Available': 575, ...}.
 
    Strategy:
      Scan all rows for cells matching our target labels, then grab the
      value from the 'Total' column (usually 2 columns to the right,
      but we find it by looking for the first numeric value after the label).
    """
    results = {}
    targets = {clean_label(t): t for t in UNIT_RATE_LABELS}
 
    for row in ws.iter_rows(values_only=False):
        cells = []
        for c in row:
            cells.append(c.value)
 
        for col_idx, val in enumerate(cells):
            if clean_label(val) in targets:
                original_label = targets[clean_label(val)]
                # Look to the right for the first numeric value (the 'Total')
                for search_col in range(col_idx + 1, min(col_idx + 5, len(cells))):
                    candidate = cells[search_col]
                    if candidate is not None:
                        try:
                            results[original_label] = float(candidate)
                            break
                        except (ValueError, TypeError):
                            continue
 
    return results
 
 
# ---------------------------------------------------------------------------
# EXTRACTION: Ops Sum
# ---------------------------------------------------------------------------
 
def extract_ops_sum(ws):
    """
    Extract rental activity rows (Rentals During Month, Walk In, etc.)
    with their associated monthly dates.
    Returns (dates, rows) where rows is a list of dicts.
 
    Strategy:
      1. Find the date row by scanning for a row with 5+ date cells.
      2. Find target rows by label matching (stripping leading whitespace).
      3. Align values with the date columns.
    """
 
    all_rows = []
    for row in ws.iter_rows(values_only=False):
        cells = []
        for c in row:
            cells.append(c.value)
        all_rows.append(cells)
 
    if not all_rows:
        return None, None
 
    # -- Find the date header row --
    date_row_idx = None
    date_start_col = None
    dates = []
 
    for idx, row in enumerate(all_rows):
        date_count = 0
        first_date_col = None
        for col_idx, val in enumerate(row):
            if is_date_value(val):
                date_count += 1
                if first_date_col is None:
                    first_date_col = col_idx
        if date_count >= 5:
            date_row_idx = idx
            date_start_col = first_date_col
            # Collect only actual date cells (skip trailing text like 'YoY Delta')
            for col_idx in range(first_date_col, len(row)):
                val = row[col_idx]
                if is_date_value(val):
                    dates.append(format_date(val))
                elif val is None:
                    break  # stop at first empty cell
                else:
                    break  # stop at non-date text
            break
 
    if date_row_idx is None:
        return None, None
 
    # -- Build a lookup of target labels (lowercase, stripped) --
    targets = {clean_label(t): t for t in OPS_SUM_LABELS}
    num_date_cols = len(dates)
 
    # -- Find the label column (usually col B = index 1) --
    # Scan a few rows after the date row to find any target label
    label_col = None
    for idx in range(date_row_idx + 1, min(date_row_idx + 40, len(all_rows))):
        row = all_rows[idx]
        for col_idx in range(min(4, len(row))):
            cell_clean = clean_label(row[col_idx])
            if cell_clean in targets:
                label_col = col_idx
                break
        if label_col is not None:
            break
 
    if label_col is None:
        return dates, None
 
    # -- Extract matching rows --
    extracted_rows = []
 
    for idx in range(date_row_idx + 1, len(all_rows)):
        row = all_rows[idx]
        raw_label = row[label_col] if label_col < len(row) else None
        if raw_label is None:
            continue
        cell_clean = clean_label(raw_label)
 
        if cell_clean in targets:
            values = []
            for d in range(num_date_cols):
                col = date_start_col + d
                val = row[col] if col < len(row) else None
                values.append(val)
 
            extracted_rows.append({
                "label": targets[cell_clean],  # use the canonical label
                "values": values,
            })
 
    return dates, extracted_rows
 
 
# ---------------------------------------------------------------------------
# EXTRACTION: Rent Roll
# ---------------------------------------------------------------------------
 
def calculate_sq_ft(size_str):
    """
    Convert a unit size string like '10X13' into its square footage (130).
 
    Handles formats like '10X13', '5x10', '05X05' (case-insensitive).
    Returns the integer product, or None if the format is unrecognized.
    """
    if size_str is None:
        return None
    # Match patterns like '10X13', '5x10', '05X05'
    match = re.match(r"(\d+)\s*[Xx]\s*(\d+)", str(size_str).strip())
    if match:
        width = int(match.group(1))
        depth = int(match.group(2))
        return width * depth
    return None
 
 
def extract_rent_roll(ws):
    """
    Extract the rent roll data from the Rent Roll sheet.
    Returns (headers, data_rows) where:
      - headers is a list of column header strings
      - data_rows is a list of lists (one per tenant/unit)
 
    Strategy:
      1. Find the header row by scanning for a row containing 'Tenant Account'
         and 'Unit #' (rather than assuming row 5).
      2. Map each expected column name to its position in that row.
      3. Read all rows below the header until we hit an empty row (col B is None).
      4. Add a calculated 'Sq Ft' column based on the 'Size' column.
    """
 
    # -- Read all rows into a list --
    all_rows = []
    for row in ws.iter_rows(values_only=False):
        cells = []
        for c in row:
            cells.append(c.value)
        all_rows.append(cells)
 
    if not all_rows:
        return None, None
 
    # -- Find the header row by looking for 'Tenant Account' --
    header_row_idx = None
    col_map = {}  # maps header name -> column index
 
    for idx, row in enumerate(all_rows):
        # Check if this row contains at least 2 of our expected headers
        row_labels = {clean_label(cell): col_idx for col_idx, cell in enumerate(row)
                      if cell is not None}
        matches = 0
        for expected in RENT_ROLL_HEADERS:
            if clean_label(expected) in row_labels:
                matches += 1
        if matches >= 3:  # found the header row
            header_row_idx = idx
            # Build the column map
            for expected in RENT_ROLL_HEADERS:
                key = clean_label(expected)
                if key in row_labels:
                    col_map[expected] = row_labels[key]
            break
 
    if header_row_idx is None:
        return None, None
 
    # -- Read data rows below the header until we hit an empty Tenant Account --
    data_rows = []
    # Figure out which column holds Tenant Account (used to detect end of data)
    tenant_col = col_map.get("Tenant Account")
 
    for idx in range(header_row_idx + 1, len(all_rows)):
        row = all_rows[idx]
 
        # Check if this row has any data at all across our mapped columns.
        # Vacant units may have a blank Tenant Account but still have a
        # Unit #, Street Rate, Status, Size, etc.  We only stop when the
        # entire row is empty across all our columns.
        has_data = False
        for header in RENT_ROLL_HEADERS:
            if header in col_map:
                col_idx = col_map[header]
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None and str(val).strip() != "":
                    has_data = True
                    break
        if not has_data:
            break
 
        # Extract each column value using the column map
        row_values = []
        for header in RENT_ROLL_HEADERS:
            if header in col_map:
                col_idx = col_map[header]
                val = row[col_idx] if col_idx < len(row) else None
                row_values.append(val)
            else:
                row_values.append(None)
 
        # Skip vacant/available units — we only want occupied units
        status_idx = RENT_ROLL_HEADERS.index("Status") if "Status" in RENT_ROLL_HEADERS else None
        if status_idx is not None and status_idx < len(row_values):
            status_val = str(row_values[status_idx] or "").strip().lower()
            if status_val == "available":
                continue
 
        # Calculate Sq Ft from the Size column
        size_idx = RENT_ROLL_HEADERS.index("Size") if "Size" in RENT_ROLL_HEADERS else None
        if size_idx is not None and size_idx < len(row_values):
            sq_ft = calculate_sq_ft(row_values[size_idx])
        else:
            sq_ft = None
        row_values.append(sq_ft)
 
        data_rows.append(row_values)
 
    # The output headers include our calculated column
    output_headers = RENT_ROLL_HEADERS + ["Sq Ft"]
 
    return output_headers, data_rows
 
 
# ---------------------------------------------------------------------------
# EXCEL OUTPUT WRITERS
# ---------------------------------------------------------------------------
 
# Reusable styles for the output workbook
HEADER_FONT = Font(bold=True)
NUMBER_FORMAT = '#,##0.00'
INTEGER_FORMAT = '#,##0'
DATE_FORMAT = 'm/d/yyyy'
 
 
def parse_date_string(date_str):
    """
    Convert a date string like 'Feb 2025' into the components needed
    for the long-format output:
      - month:  integer (2)
      - year:   integer (2025)
      - period: datetime object for the 1st of that month (2/1/2025)
 
    Returns (month, year, period_date) or (None, None, None) on failure.
    """
    try:
        dt = datetime.strptime(date_str, "%b %Y")
        return dt.month, dt.year, dt  # dt is already the 1st of the month
    except ValueError:
        return None, None, None
 
 
def write_rolling_is_tab(out_wb, source_file, prop_num, dates, rows, property_name=""):
    """
    Write the Rolling IS data in LONG format — one row per line item per month.
 
    This is the format used for importing into pro forma builder templates.
 
    Columns: Property Name, line_item, Month, Year, Period, Amount
 
    Where:
      - Month  = integer month number (e.g. 2 for February)
      - Year   = four-digit year (e.g. 2025)
      - Period = date as m/d/yyyy representing the 1st of the month (e.g. 2/1/2025)
      - Amount = the dollar value for that line item in that month
    """
    ws = out_wb.create_sheet(title="Rolling IS")
 
    # -- Row 1-2: metadata --
    ws["A1"] = "Source File"
    ws["B1"] = source_file
    ws["A2"] = "Property Number"
    ws["B2"] = prop_num
    ws["A1"].font = HEADER_FONT
    ws["A2"].font = HEADER_FONT
 
    # -- Row 4: column headers --
    header_row = 4
    headers = ["Property Name", "line_item", "Month", "Year", "Period", "Amount"]
    for col_idx, header in enumerate(headers):
        ws.cell(row=header_row, column=col_idx + 1, value=header).font = HEADER_FONT
 
    # -- Row 5+: one row per line item per month --
    data_row = header_row + 1
 
    for row_data in rows:
        for i, date_str in enumerate(dates):
            month, year, period_date = parse_date_string(date_str)
            amount = row_data["values"][i] if i < len(row_data["values"]) else 0
 
            line_item = row_data["label"]
            # Col A: Property Name (only where line_item is not blank)
            ws.cell(row=data_row, column=1, value=property_name if line_item else "")
            ws.cell(row=data_row, column=2, value=line_item)
            ws.cell(row=data_row, column=3, value=month)
            ws.cell(row=data_row, column=4, value=year)
 
            period_cell = ws.cell(row=data_row, column=5, value=period_date)
            period_cell.number_format = DATE_FORMAT
 
            amount_cell = ws.cell(row=data_row, column=6, value=amount if amount is not None else 0)
            amount_cell.number_format = NUMBER_FORMAT
 
            data_row += 1
 
    # -- Column widths --
    ws.column_dimensions["A"].width = 20   # Property Name
    ws.column_dimensions["B"].width = 40   # line_item
    ws.column_dimensions["C"].width = 8    # Month
    ws.column_dimensions["D"].width = 8    # Year
    ws.column_dimensions["E"].width = 12   # Period
    ws.column_dimensions["F"].width = 14   # Amount
 
    return ws
 
 
def write_unit_rate_tab(out_wb, source_file, prop_num, metrics):
    """
    Write the Unit Rate metrics as a simple two-column table.
 
    Layout:
      Row 1-2: metadata
      Row 4: header — "Metric", "Value"
      Row 5+: one row per metric
    """
    ws = out_wb.create_sheet(title="Unit Rate")
 
    # -- Metadata --
    ws["A1"] = "Source File"
    ws["B1"] = source_file
    ws["A2"] = "Property Number"
    ws["B2"] = prop_num
    ws["A1"].font = HEADER_FONT
    ws["A2"].font = HEADER_FONT
 
    # -- Header --
    ws.cell(row=4, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=4, column=2, value="Value").font = HEADER_FONT
 
    # -- Data rows (in the order defined in UNIT_RATE_LABELS) --
    row_num = 5
    for label in UNIT_RATE_LABELS:
        if label in metrics:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=metrics[label])
            cell.number_format = INTEGER_FORMAT
            row_num += 1
 
    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
 
    return ws
 
 
def write_ops_sum_tab(out_wb, source_file, prop_num, dates, rows):
    """
    Write the Ops Sum data in LONG format — one row per metric per month.
    Same column structure as Rolling IS for consistency.
 
    Columns: metric, Month, Year, Period, Value
    """
    ws = out_wb.create_sheet(title="Ops Sum")
 
    # -- Metadata --
    ws["A1"] = "Source File"
    ws["B1"] = source_file
    ws["A2"] = "Property Number"
    ws["B2"] = prop_num
    ws["A1"].font = HEADER_FONT
    ws["A2"].font = HEADER_FONT
 
    # -- Header --
    header_row = 4
    headers = ["metric", "Month", "Year", "Period", "Value"]
    for col_idx, header in enumerate(headers):
        ws.cell(row=header_row, column=col_idx + 1, value=header).font = HEADER_FONT
 
    # -- Data rows: one row per metric per month --
    data_row = header_row + 1
 
    for row_data in rows:
        for i, date_str in enumerate(dates):
            month, year, period_date = parse_date_string(date_str)
            value = row_data["values"][i] if i < len(row_data["values"]) else 0
 
            ws.cell(row=data_row, column=1, value=row_data["label"])
            ws.cell(row=data_row, column=2, value=month)
            ws.cell(row=data_row, column=3, value=year)
 
            period_cell = ws.cell(row=data_row, column=4, value=period_date)
            period_cell.number_format = DATE_FORMAT
 
            value_cell = ws.cell(row=data_row, column=5, value=value if value is not None else 0)
            value_cell.number_format = INTEGER_FORMAT
 
            data_row += 1
 
    # Column widths
    ws.column_dimensions["A"].width = 26   # metric
    ws.column_dimensions["B"].width = 8    # Month
    ws.column_dimensions["C"].width = 8    # Year
    ws.column_dimensions["D"].width = 12   # Period
    ws.column_dimensions["E"].width = 10   # Value
 
    return ws
 
 
def write_rent_roll_tab(out_wb, source_file, prop_num, headers, data_rows):
    """
    Write the Rent Roll data as a flat table.
 
    Layout:
      Row 1-2: metadata (source file, property number)
      Row 4:   column headers (Tenant Account, Unit #, ... Size, Type, Sq Ft)
      Row 5+:  one row per unit/tenant
 
    The 'Sq Ft' column is a calculated field based on the 'Size' column.
    For example, Size='10X13' produces Sq Ft=130.
    """
    ws = out_wb.create_sheet(title="Rent Roll")
 
    # -- Metadata --
    ws["A1"] = "Source File"
    ws["B1"] = source_file
    ws["A2"] = "Property Number"
    ws["B2"] = prop_num
    ws["A1"].font = HEADER_FONT
    ws["A2"].font = HEADER_FONT
 
    # -- Header row --
    header_row = 4
    for col_idx, header in enumerate(headers):
        ws.cell(row=header_row, column=col_idx + 1, value=header).font = HEADER_FONT
 
    # -- Data rows --
    # Define which columns get special formatting
    date_columns = {"Move-In Date", "Paid-Thru Date"}
    money_columns = {"Rent Rate", "Street Rate"}
    integer_columns = {"Sq Ft"}
 
    for row_idx, row_values in enumerate(data_rows):
        data_row = header_row + 1 + row_idx
 
        for col_idx, val in enumerate(row_values):
            cell = ws.cell(row=data_row, column=col_idx + 1, value=val)
 
            # Apply formatting based on the column header
            col_name = headers[col_idx] if col_idx < len(headers) else ""
            if col_name in date_columns and isinstance(val, datetime):
                cell.number_format = DATE_FORMAT
            elif col_name in money_columns:
                cell.number_format = NUMBER_FORMAT
            elif col_name in integer_columns and val is not None:
                cell.number_format = INTEGER_FORMAT
 
    # -- Column widths --
    col_widths = {
        "Tenant Account": 16,
        "Unit #": 10,
        "Move-In Date": 13,
        "Rent Rate": 12,
        "Street Rate": 12,
        "Paid-Thru Date": 15,
        "Status": 10,
        "Size": 8,
        "Type": 8,
        "Sq Ft": 8,
    }
    for col_idx, header in enumerate(headers):
        col_letter = chr(ord("A") + col_idx) if col_idx < 26 else "A"
        ws.column_dimensions[col_letter].width = col_widths.get(header, 12)
 
    return ws
 
 
def write_log_tab(out_wb, log_entries):
    """
    Write the processing log as a tab in the output workbook.
 
    Columns: Timestamp, Sheet, Status, Message
    """
    ws = out_wb.create_sheet(title="Processing Log")
 
    headers = ["Timestamp", "Sheet", "Status", "Message"]
    for col_idx, header in enumerate(headers):
        ws.cell(row=1, column=col_idx + 1, value=header).font = HEADER_FONT
 
    for row_idx, entry in enumerate(log_entries):
        for col_idx, val in enumerate(entry):
            ws.cell(row=row_idx + 2, column=col_idx + 1, value=val)
 
    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 55
 
    return ws
 
 
# ---------------------------------------------------------------------------
# MAIN PROCESSING LOOP
# ---------------------------------------------------------------------------
 
def process_file(filepath):
    """
    Process a single Excel file:
      1. Open and extract data from each sheet
      2. Prompt the user for a property name
      3. Write one output .xlsx workbook with all extracted tabs
 
    Returns the output filepath (or None if the file could not be opened).
    """
    filename = os.path.basename(filepath)
    log_entries = []  # log entries scoped to this file only
 
    print(f"\n{'=' * 50}")
    print(f"  File: {filename}")
    print(f"{'=' * 50}")
 
    # -- Open the input workbook --
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        msg = f"Could not open file: {e}"
        print(f"  ERROR: {msg}")
        return None
 
    # -- Extract from each sheet --
 
    # Rolling IS
    rolling_is_data = None
    sheet_name, ws = find_sheet_by_prefix(wb, SHEET_PREFIXES["rolling_is"])
    if ws is None:
        msg = f"Sheet with prefix '{SHEET_PREFIXES['rolling_is']}' not found"
        print(f"  WARNING: {msg}")
        log_entries.append([datetime.now().isoformat(), SHEET_PREFIXES["rolling_is"],
                           "WARNING", msg])
    else:
        prop_num = extract_property_number(sheet_name, SHEET_PREFIXES["rolling_is"])
        dates, rows = extract_rolling_is(ws)
 
        if dates is None:
            msg = "Could not find date header row"
            print(f"  WARNING: {msg} in {sheet_name}")
            log_entries.append([datetime.now().isoformat(), sheet_name, "WARNING", msg])
        elif rows is None:
            msg = f"Could not find '{ROLLING_IS_START_LABEL}' label"
            print(f"  WARNING: {msg} in {sheet_name}")
            log_entries.append([datetime.now().isoformat(), sheet_name, "WARNING", msg])
        else:
            rolling_is_data = {"prop_num": prop_num, "dates": dates, "rows": rows}
            msg = f"Extracted {len(rows)} line items x {len(dates)} months"
            print(f"  OK: Rolling IS -> {msg}")
            log_entries.append([datetime.now().isoformat(), sheet_name, "OK", msg])
 
    # Unit Rate
    unit_rate_data = None
    sheet_name, ws = find_sheet_by_prefix(wb, SHEET_PREFIXES["unit_rate"])
    if ws is None:
        msg = f"Sheet with prefix '{SHEET_PREFIXES['unit_rate']}' not found"
        print(f"  WARNING: {msg}")
        log_entries.append([datetime.now().isoformat(), SHEET_PREFIXES["unit_rate"],
                           "WARNING", msg])
    else:
        prop_num = extract_property_number(sheet_name, SHEET_PREFIXES["unit_rate"])
        metrics = extract_unit_rate(ws)
 
        if not metrics:
            msg = "No matching metrics found"
            print(f"  WARNING: {msg} in {sheet_name}")
            log_entries.append([datetime.now().isoformat(), sheet_name, "WARNING", msg])
        else:
            found = set(metrics.keys())
            expected = set(UNIT_RATE_LABELS)
            missing = expected - found
            if missing:
                msg = f"Missing labels: {', '.join(missing)}"
                print(f"  WARNING: {msg}")
                log_entries.append([datetime.now().isoformat(), sheet_name,
                                   "WARNING", msg])
 
            unit_rate_data = {"prop_num": prop_num, "metrics": metrics}
            msg = f"Extracted {len(metrics)} metrics"
            print(f"  OK: Unit Rate -> {msg}")
            log_entries.append([datetime.now().isoformat(), sheet_name, "OK", msg])
 
    # Ops Sum
    ops_sum_data = None
    sheet_name, ws = find_sheet_by_prefix(wb, SHEET_PREFIXES["ops_sum"])
    if ws is None:
        msg = f"Sheet with prefix '{SHEET_PREFIXES['ops_sum']}' not found"
        print(f"  WARNING: {msg}")
        log_entries.append([datetime.now().isoformat(), SHEET_PREFIXES["ops_sum"],
                           "WARNING", msg])
    else:
        prop_num = extract_property_number(sheet_name, SHEET_PREFIXES["ops_sum"])
        dates, rows = extract_ops_sum(ws)
 
        if dates is None:
            msg = "Could not find date header row"
            print(f"  WARNING: {msg} in {sheet_name}")
            log_entries.append([datetime.now().isoformat(), sheet_name, "WARNING", msg])
        elif rows is None:
            msg = "Could not find label column for rental activity rows"
            print(f"  WARNING: {msg} in {sheet_name}")
            log_entries.append([datetime.now().isoformat(), sheet_name, "WARNING", msg])
        else:
            found_labels = {r["label"] for r in rows}
            expected_labels = set(OPS_SUM_LABELS)
            missing = expected_labels - found_labels
            if missing:
                msg = f"Missing rows: {', '.join(missing)}"
                print(f"  WARNING: {msg}")
                log_entries.append([datetime.now().isoformat(), sheet_name,
                                   "WARNING", msg])
 
            ops_sum_data = {"prop_num": prop_num, "dates": dates, "rows": rows}
            msg = f"Extracted {len(rows)} metrics x {len(dates)} months"
            print(f"  OK: Ops Sum -> {msg}")
            log_entries.append([datetime.now().isoformat(), sheet_name, "OK", msg])
 
    # Rent Roll
    rent_roll_data = None
    sheet_name, ws = find_sheet_by_prefix(wb, SHEET_PREFIXES["rent_roll"])
    if ws is None:
        msg = f"Sheet with prefix '{SHEET_PREFIXES['rent_roll']}' not found"
        print(f"  WARNING: {msg}")
        log_entries.append([datetime.now().isoformat(), SHEET_PREFIXES["rent_roll"],
                           "WARNING", msg])
    else:
        prop_num = extract_property_number(sheet_name, SHEET_PREFIXES["rent_roll"])
        headers, data_rows = extract_rent_roll(ws)
 
        if headers is None:
            msg = "Could not find rent roll header row"
            print(f"  WARNING: {msg} in {sheet_name}")
            log_entries.append([datetime.now().isoformat(), sheet_name, "WARNING", msg])
        elif not data_rows:
            msg = "No tenant/unit rows found"
            print(f"  WARNING: {msg} in {sheet_name}")
            log_entries.append([datetime.now().isoformat(), sheet_name, "WARNING", msg])
        else:
            rent_roll_data = {
                "prop_num": prop_num,
                "headers": headers,
                "data_rows": data_rows,
            }
            msg = f"Extracted {len(data_rows)} units x {len(headers)} columns"
            print(f"  OK: Rent Roll -> {msg}")
            log_entries.append([datetime.now().isoformat(), sheet_name, "OK", msg])
 
    wb.close()
 
    # -- Prompt the user for a property name --
    property_name = prompt_for_property_name(filename)
    safe_name = make_safe_filename(property_name)
    output_filename = f"{safe_name}_datapack.xlsx"
    output_path = os.path.join("output", output_filename)
 
    # -- Build the output workbook --
    out_wb = Workbook()
    # Remove the default "Sheet" that openpyxl creates automatically
    out_wb.remove(out_wb.active)
 
    # Write each tab (only if data was successfully extracted)
    if rolling_is_data:
        write_rolling_is_tab(out_wb, filename, rolling_is_data["prop_num"],
                             rolling_is_data["dates"], rolling_is_data["rows"],
                             property_name=property_name)
    else:
        print("  (no Rolling IS data to write)")
 
    if unit_rate_data:
        write_unit_rate_tab(out_wb, filename, unit_rate_data["prop_num"],
                            unit_rate_data["metrics"])
    else:
        print("  (no Unit Rate data to write)")
 
    if ops_sum_data:
        write_ops_sum_tab(out_wb, filename, ops_sum_data["prop_num"],
                          ops_sum_data["dates"], ops_sum_data["rows"])
    else:
        print("  (no Ops Sum data to write)")
 
    if rent_roll_data:
        write_rent_roll_tab(out_wb, filename, rent_roll_data["prop_num"],
                            rent_roll_data["headers"], rent_roll_data["data_rows"])
    else:
        print("  (no Rent Roll data to write)")
 
    # Always include the processing log tab
    write_log_tab(out_wb, log_entries)
 
    # -- Save the output workbook --
    out_wb.save(output_path)
    print(f"\n  -> Saved: {output_path}")
 
    return output_path
 
 
def main():
    """Main entry point."""
    print("=" * 60)
    print("  EXR Owner Financials Extractor  v2.0")
    print("=" * 60)
 
    # Set up folders
    setup_folders()
 
    # Find files
    files = find_xlsx_files("input")
    if not files:
        print("\n  No .xlsx files found in the 'input' folder.")
        print("  Drop your owner financial files there and run again.")
        print()
        sys.exit(0)
 
    print(f"\n  Found {len(files)} file(s) to process.")
 
    # Process each file independently (one output workbook per input file)
    results = []
    for filepath in files:
        output_path = process_file(filepath)
        if output_path:
            results.append(output_path)
 
    # Final summary
    print(f"\n{'=' * 60}")
    print(f"  Done!  {len(results)} output file(s) created in the 'output' folder:")
    for path in results:
        print(f"    {os.path.basename(path)}")
    print("=" * 60)
 
 
if __name__ == "__main__":
    main()
